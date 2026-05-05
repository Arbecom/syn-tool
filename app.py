#!/usr/bin/env python3
"""
Synology NAS Storage Management Tool
- Standalone: Python 3.9 (DSM Package Center, works on DSM 7.0+)
- Docker: see Dockerfile (requires DSM 7.2+)

Scan NAS shares, manage billing Excel, track version history.
"""

import os
import re
import json
import subprocess
import datetime
import io
import threading
import hashlib
import http.cookiejar
import urllib.request
import urllib.parse
from pathlib import Path
from typing import Optional
import secrets
from datetime import timedelta
from flask import Flask, request, jsonify, send_file, send_from_directory, Response, session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder="static")
app.permanent_session_lifetime = timedelta(days=30)

# ---------------------------------------------------------------------------
# Paths & defaults
# ---------------------------------------------------------------------------

DATA_DIR         = Path(os.environ.get("DATA_DIR", "./data"))
_branding_env    = os.environ.get("BRANDING_DIR", "")
BRANDING_DIR     = Path(_branding_env) if _branding_env else None
UPLOADS_DIR      = DATA_DIR / "uploads"
EDITS_DIR        = DATA_DIR / "edits"
MAPPINGS_DIR     = DATA_DIR / "mappings_history"
CURRENT_FILE          = DATA_DIR / "current.json"
CONFIG_FILE           = DATA_DIR / "config.json"
MAPPINGS_FILE         = DATA_DIR / "mappings.json"
APPARENT_SIZES_FILE   = DATA_DIR / "apparent_sizes.json"

_scan_lock = threading.Lock()   # Only one share scan may run at a time

DEFAULT_CONFIG = {
    "share_paths":       ["/volume1"],
    "exclude_shares":    ["@eaDir", "@sharebin", "#recycle", "@tmp", "homes"],
    "upload_retention":  10,
    "edit_retention":    10,
    "mailbox_gb":        10,    # GB included per mailbox (Te factureren = Gebruikte - mailbox_gb * Mailboxen)
    "auth_enabled":      True,
    "auth_username":     "admin",
    "auth_password":     "admin",
    "dsm_host":          "localhost",
    "dsm_port":          3333,
    "dsm_user":          "",
    "dsm_password":      "",
}

# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------

def ensure_dirs():
    for d in [DATA_DIR, UPLOADS_DIR, EDITS_DIR, MAPPINGS_DIR]:
        d.mkdir(parents=True, exist_ok=True)
    key_file = DATA_DIR / ".secret_key"
    if not key_file.exists():
        key_file.write_text(secrets.token_hex(32))
    app.secret_key = key_file.read_text().strip()


def load_config() -> dict:
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE) as f:
            cfg = json.load(f)
        for k, v in DEFAULT_CONFIG.items():
            cfg.setdefault(k, v)
    else:
        cfg = dict(DEFAULT_CONFIG)
    return _apply_env_overrides(cfg)


def save_config(cfg: dict):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)


def load_apparent_cache() -> dict:
    if APPARENT_SIZES_FILE.exists():
        try:
            with open(APPARENT_SIZES_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_apparent_cache(cache: dict):
    try:
        with open(APPARENT_SIZES_FILE, "w") as f:
            json.dump(cache, f)
    except Exception:
        pass


_PBKDF2_ITERS = 50000

def _encrypt_credential(plaintext: str) -> str:
    """Encrypt a credential with the app secret key. Stored as 'enc:<base64>'."""
    import base64
    data   = plaintext.encode('utf-8')
    iv     = secrets.token_bytes(16)
    secret = (DATA_DIR / ".secret_key").read_text().strip().encode()
    key    = hashlib.pbkdf2_hmac('sha256', secret, iv, 1000, dklen=len(data))
    return 'enc:' + base64.b64encode(iv + bytes(b ^ k for b, k in zip(data, key))).decode()

def _decrypt_credential(stored: str) -> str:
    """Decrypt a credential encrypted by _encrypt_credential. Returns plaintext."""
    import base64
    if not stored.startswith('enc:'):
        return stored  # legacy plaintext — transparently supported
    try:
        raw     = base64.b64decode(stored[4:])
        iv, enc = raw[:16], raw[16:]
        secret  = (DATA_DIR / ".secret_key").read_text().strip().encode()
        key     = hashlib.pbkdf2_hmac('sha256', secret, iv, 1000, dklen=len(enc))
        return bytes(b ^ k for b, k in zip(enc, key)).decode('utf-8')
    except Exception:
        return ''

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), _PBKDF2_ITERS)
    return f"pbkdf2:sha256:{_PBKDF2_ITERS}:{salt}:{h.hex()}"

def verify_password(password: str, stored: str) -> bool:
    if not stored.startswith('pbkdf2:'):
        return password == stored  # legacy plaintext
    try:
        _, algo, iters, salt, expected = stored.split(':', 4)
        h = hashlib.pbkdf2_hmac(algo, password.encode('utf-8'), salt.encode('utf-8'), int(iters))
        return h.hex() == expected
    except Exception:
        return False


def _read_secret(env_name: str) -> Optional[str]:
    """Read a sensitive value from (in priority order):
    1. {env_name}_FILE env var → path to a file (Docker secrets pattern)
    2. {env_name} env var directly
    3. /run/secrets/<lowercase-name-without-SYNTOOL_prefix>  (Docker swarm default)
    """
    file_path = os.environ.get(f"{env_name}_FILE")
    if file_path:
        try:
            return Path(file_path).read_text().strip() or None
        except OSError:
            pass
    val = os.environ.get(env_name)
    if val:
        return val
    secret_name = env_name.replace("SYNTOOL_", "").lower()
    secret_path = Path("/run/secrets") / secret_name
    if secret_path.exists():
        try:
            return secret_path.read_text().strip() or None
        except OSError:
            pass
    return None


def _apply_env_overrides(cfg: dict) -> dict:
    """Apply SYNTOOL_* environment variables (and Docker secrets) on top of config.
    Non-secret settings are applied on every call. Passwords are hashed/encrypted
    once and written back to config — a sentinel detects when the env value changes."""
    for env_key, cfg_key, conv in [
        ("SYNTOOL_AUTH_USERNAME", "auth_username", str),
        ("SYNTOOL_AUTH_ENABLED",  "auth_enabled",  lambda v: v.strip().lower() in ("true", "1", "yes")),
        ("SYNTOOL_DSM_HOST",      "dsm_host",      str),
        ("SYNTOOL_DSM_PORT",      "dsm_port",      int),
        ("SYNTOOL_DSM_USER",      "dsm_user",      str),
        ("SYNTOOL_MAILBOX_GB",    "mailbox_gb",    float),
    ]:
        val = os.environ.get(env_key)
        if val is not None:
            try:
                cfg[cfg_key] = conv(val)
            except (ValueError, TypeError):
                pass

    auth_pw = _read_secret("SYNTOOL_AUTH_PASSWORD")
    if auth_pw:
        sentinel = hashlib.sha256(auth_pw.encode()).hexdigest()
        if cfg.get("_env_auth_pw_sentinel") != sentinel:
            cfg["auth_password"]         = hash_password(auth_pw)
            cfg["_env_auth_pw_sentinel"] = sentinel
            save_config(cfg)

    dsm_pw = _read_secret("SYNTOOL_DSM_PASSWORD")
    if dsm_pw:
        sentinel = hashlib.sha256(dsm_pw.encode()).hexdigest()
        if cfg.get("_env_dsm_pw_sentinel") != sentinel:
            cfg["dsm_password"]         = _encrypt_credential(dsm_pw)
            cfg["_env_dsm_pw_sentinel"] = sentinel
            save_config(cfg)

    return cfg


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.before_request
def check_auth():
    if not request.path.startswith('/api/'):
        return
    if request.path.startswith('/api/auth/'):
        return
    cfg = load_config()
    if not cfg.get('auth_enabled', True):
        return
    if not session.get('authenticated'):
        return jsonify({'error': 'Unauthorized'}), 401


@app.route('/api/auth/status')
def auth_status():
    cfg = load_config()
    enabled = cfg.get('auth_enabled', True)
    return jsonify({
        'authenticated': not enabled or bool(session.get('authenticated')),
        'auth_enabled': enabled,
    })


@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    cfg = load_config()
    if not cfg.get('auth_enabled', True):
        return jsonify({'ok': True})
    data = request.get_json() or {}
    username = data.get('username', '')
    password = data.get('password', '')
    stored_pw = cfg.get('auth_password', 'admin')
    if username == cfg.get('auth_username', 'admin') and verify_password(password, stored_pw):
        if not stored_pw.startswith('pbkdf2:'):
            cfg['auth_password'] = hash_password(password)
            save_config(cfg)
        session.permanent = True
        session['authenticated'] = True
        return jsonify({'ok': True})
    return jsonify({'error': 'Invalid credentials'}), 401


@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'ok': True})


# ---------------------------------------------------------------------------
# Customer → share mappings
# ---------------------------------------------------------------------------

def load_mappings() -> dict:
    if MAPPINGS_FILE.exists():
        with open(MAPPINGS_FILE) as f:
            return json.load(f)
    return {"key_col": None, "map": {}}


def _snapshot_mappings():
    if not MAPPINGS_FILE.exists():
        return
    with open(MAPPINGS_FILE) as f:
        old = json.load(f)
    old["_snap_at"] = datetime.datetime.now().isoformat()
    with open(MAPPINGS_DIR / f"{_ts()}.json", "w") as f:
        json.dump(old, f, indent=2)


def save_mappings(data: dict):
    _snapshot_mappings()
    with open(MAPPINGS_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)
    cfg = load_config()
    _apply_retention(MAPPINGS_DIR, "*.json", cfg.get("edit_retention", 10))


def detect_key_col(headers: list) -> Optional[str]:
    """Customer name is the key — more reliably filled than contract number."""
    return (
        _find_col(headers, "klant", "naam", "customer", "name") or
        _find_col(headers, "contract")
    )


def compute_mapping_diff(headers: list, rows: list, mappings: dict) -> dict:
    """
    Apply stored mappings to rows and return a diff describing what changed.
    Mutates rows in-place by setting row['_share'] where a mapping exists.
    """
    key_col  = detect_key_col(headers)
    name_col = _find_col(headers, "klant", "naam", "customer", "name")
    old_map  = mappings.get("map", {})

    applied, new_rows, changed = [], [], []
    seen_keys = set()

    for row in rows:
        display_name = str(row.get(key_col, "")).strip() if key_col else ""
        if not display_name:
            continue
        key = display_name.lower()          # lowercase for comparison / storage
        seen_keys.add(key)
        # Prefer the dedicated name column for display; fall back to key column value
        name = str(row.get(name_col, display_name)).strip() if (name_col and name_col != key_col) else display_name

        if key in old_map:
            entry = old_map[key]
            row["_share"] = entry.get("share")
            old_name = entry.get("name", "")
            if old_name and name and old_name.lower() != name.lower():
                changed.append({"key": key, "old_name": old_name, "new_name": name,
                                 "share": entry.get("share")})
            applied.append({"key": key, "name": name, "share": entry.get("share")})
        else:
            new_rows.append({"key": key, "name": name})

    removed = [
        {"key": k, "name": v.get("name", k), "share": v.get("share")}
        for k, v in old_map.items() if k not in seen_keys
    ]

    return {
        "key_col":  key_col,
        "applied":  applied,
        "new":      new_rows,
        "removed":  removed,
        "changed":  changed,
        "has_diff": bool(new_rows or removed or changed),
    }


# ---------------------------------------------------------------------------
# NAS share scanning
# ---------------------------------------------------------------------------

def _get_dsm_analyzer_sizes(cfg: dict):
    """Fetch share sizes from DSM Storage Analyzer via HTTP API.
    Authenticates, lists Storage Analyzer reports, fetches each report HTML,
    and parses the embedded JS array: ['ShareName', 'volume_N', 'size_bytes', ...]
    Returns (sizes, dates) where both are {share_name: value}:
      sizes — {share_name: size_bytes}
      dates — {share_name: 'YYYY-MM-DD'} date of the most recent report scanned"""
    host     = str(cfg.get('dsm_host', '')).strip()
    port     = cfg.get('dsm_port', 3333)
    user     = str(cfg.get('dsm_user', '')).strip()
    password = _decrypt_credential(str(cfg.get('dsm_password', '')).strip())

    if not host or not user or not password:
        return {}, {}

    base_url = f"http://{host}:{port}"
    cj       = http.cookiejar.CookieJar()
    opener   = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))

    login_data = urllib.parse.urlencode({
        'api': 'SYNO.API.Auth', 'version': '3', 'method': 'login',
        'account': user, 'passwd': password,
        'session': 'SynologyTool', 'format': 'cookie',
    })
    try:
        resp = opener.open(f"{base_url}/webapi/entry.cgi", login_data.encode(), timeout=10)
        if not json.loads(resp.read()).get('success'):
            return {}, {}
    except Exception:
        return {}, {}

    list_data = urllib.parse.urlencode({
        'api': 'SYNO.Core.Report', 'version': '1', 'method': 'list',
    })
    try:
        resp    = opener.open(f"{base_url}/webapi/entry.cgi", list_data.encode(), timeout=15)
        reports = json.loads(resp.read()).get('data', {}).get('reports', [])
    except Exception:
        return {}, {}

    sizes   = {}
    dates   = {}
    pattern = re.compile(r"\[\s*'([^']+)'\s*,\s*'volume_\d+'\s*,\s*'(\d+)'")

    for report in reports:
        link = report.get('link', '')
        if not link or report.get('status') != 'success':
            continue
        # Extract scan date from link path: /dar/Report Name/2026-04-28_08-55-48/report.html
        scan_date = ''
        parts = link.rstrip('/').split('/')
        if len(parts) >= 2:
            ts = parts[-2]           # e.g. '2026-04-28_08-55-48'
            scan_date = ts[:10]      # 'YYYY-MM-DD'
        try:
            resp = opener.open(base_url + urllib.parse.quote(link), timeout=30)
            html = resp.read().decode('utf-8', errors='replace')
        except Exception:
            continue
        for m in pattern.finditer(html):
            name, sz = m.group(1), int(m.group(2))
            if sz > sizes.get(name, 0):
                sizes[name] = sz
                dates[name] = scan_date

    return sizes, dates


def _btrfs_sizes_for_paths(paths: list, base: str):
    """Return (sizes, du_needed) where sizes is {path: size_bytes} and du_needed is
    the set of paths where level-1 > level-0, meaning nested subvolumes cause btrfs
    to undercount vs the apparent (File Station) size — those need du --apparent-size."""
    if not paths:
        return {}, set()

    path_to_id: dict = {}
    id_lock = threading.Lock()

    def fetch_id(path):
        try:
            r = subprocess.run(
                ["btrfs", "subvolume", "show", path],
                capture_output=True, text=True, timeout=10
            )
            if r.returncode == 0:
                for line in r.stdout.split('\n'):
                    if 'Subvolume ID:' in line:
                        with id_lock:
                            path_to_id[path] = line.split()[-1].strip()
                        break
        except Exception:
            pass

    id_threads = [threading.Thread(target=fetch_id, args=(p,), daemon=True) for p in paths]
    for t in id_threads:
        t.start()
    for t in id_threads:
        t.join()

    if not path_to_id:
        return {}, set()

    # Try direct btrfs qgroup first (standalone Python — already in host namespace).
    # Fall back to nsenter (Docker — bind mounts can't answer qgroup queries, so we
    # enter PID 1's mount namespace where /volumeN is the real btrfs mount).
    rq: Optional[subprocess.CompletedProcess] = None
    for cmd in [
        ["btrfs", "qgroup", "show", "--raw", base],
        ["nsenter", "--target", "1", "--mount", "--", "btrfs", "qgroup", "show", "--raw", base],
    ]:
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            if result.returncode == 0:
                rq = result
                break
        except Exception:
            continue

    if rq is None:
        return {}, set()

    id_to_path = {sid: path for path, sid in path_to_id.items()}
    level0: dict = {}
    level1: dict = {}
    for line in rq.stdout.split('\n'):
        parts = line.split()
        if len(parts) >= 2 and '/' in parts[0]:
            qlevel, sid = parts[0].split('/', 1)
            if sid in id_to_path:
                p = id_to_path[sid]
                rfer = int(parts[1])
                if qlevel == '1':
                    level1[p] = rfer
                elif qlevel == '0':
                    level0[p] = rfer

    sizes = {**level0, **level1}
    du_needed = {p for p in sizes if level1.get(p, 0) > level0.get(p, 0)}
    return sizes, du_needed


def bytes_to_gb(b: int) -> float:
    return round(b / (1024 ** 3), 2)


def human_size(b: int) -> str:
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if b < 1024:
            return f"{b:.1f} {unit}"
        b /= 1024
    return f"{b:.1f} PB"


@app.route("/api/shares/stream")
def api_shares_stream():
    """SSE endpoint: scans shares in parallel, streams results as they finish.
    Only one scan runs at a time (_scan_lock). Sends keepalive comments every 0.5 s.
    Writes partial cache after every completed share so refreshes see partial results."""
    cfg = load_config()

    def generate():
        # Reject concurrent scans immediately
        if not _scan_lock.acquire(blocking=False):
            yield f'data: {json.dumps({"type":"busy"})}\n\n'
            return

        try:
            apparent_cache = load_apparent_cache()
            shares_to_scan = []
            volumes = {}

            for base in cfg["share_paths"]:
                if not os.path.isdir(base):
                    yield f'data: {json.dumps({"type":"error","message":f"Path not found: {base}"})}\n\n'
                    continue
                try:
                    st = os.statvfs(base)
                    volumes[base] = {
                        "total_bytes": st.f_blocks * st.f_frsize,
                        "free_bytes":  st.f_bfree  * st.f_frsize,
                        "used_bytes":  (st.f_blocks - st.f_bfree) * st.f_frsize,
                        "total_human": human_size(st.f_blocks * st.f_frsize),
                        "free_human":  human_size(st.f_bfree  * st.f_frsize),
                    }
                except OSError:
                    pass
                try:
                    with os.scandir(base) as it:
                        for e in it:
                            if not e.is_dir(follow_symlinks=False):
                                continue
                            n = e.name
                            if n in cfg["exclude_shares"] or n.startswith(("@", "#")):
                                continue
                            shares_to_scan.append((base, e.path, n))
                except PermissionError:
                    yield f'data: {json.dumps({"type":"error","message":f"Permission denied: {base}"})}\n\n'
                except OSError as ex:
                    yield f'data: {json.dumps({"type":"error","message":str(ex)})}\n\n'

            total = len(shares_to_scan)

            discovered = [{"name": n, "path": p, "base": b} for b, p, n in shares_to_scan]
            yield f'data: {json.dumps({"type":"discovered","shares":discovered,"volumes":volumes})}\n\n'

            if not total:
                yield f'data: {json.dumps({"type":"done","volumes":volumes})}\n\n'
                return

            # Fetch Storage Analyzer sizes via DSM HTTP API (accurate, bypasses ACLs).
            # Shares covered by analyzer skip du entirely; others fall back to du/btrfs.
            analyzer_sizes: dict = {}
            analyzer_dates: dict = {}
            if cfg.get('dsm_user') and cfg.get('dsm_password'):
                try:
                    analyzer_sizes, analyzer_dates = _get_dsm_analyzer_sizes(cfg)
                except Exception:
                    pass

            btrfs_sizes: dict = {}
            for base in set(b for b, p, n in shares_to_scan):
                paths = [p for b2, p, n in shares_to_scan if b2 == base]
                sz, _ = _btrfs_sizes_for_paths(paths, base)
                btrfs_sizes.update(sz)

            # Emit every share immediately — analyzer sizes are accurate; others
            # fall back to btrfs qgroup or last-known cache value.
            all_shares_collected: list = []

            for base, path, name in shares_to_scan:
                analyzer_bytes = analyzer_sizes.get(name, 0)
                if analyzer_bytes > 0:
                    sb = analyzer_bytes
                    apparent_cache[path] = sb
                else:
                    sb = apparent_cache.get(path, btrfs_sizes.get(path, 0))

                share = {
                    "name": name, "path": path, "base": base,
                    "size_bytes": sb, "size_gb": bytes_to_gb(sb),
                    "size_human": human_size(sb), "pending": False,
                    "analyzer_date": analyzer_dates.get(name, ""),
                }
                all_shares_collected.append(share)
                yield f'data: {json.dumps({"type":"share","share":share})}\n\n'

            if analyzer_sizes:
                save_apparent_cache(apparent_cache)

            try:
                with open(DATA_DIR / "shares_cache.json", "w") as f:
                    json.dump({"shares": all_shares_collected, "volumes": volumes,
                               "scanned_at": datetime.datetime.now().isoformat(),
                               "partial": False}, f, default=str)
            except Exception:
                pass

            yield f'data: {json.dumps({"type":"done","volumes":volumes})}\n\n'
            yield f'data: {json.dumps({"type":"all_done"})}\n\n'

        finally:
            _scan_lock.release()

    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )

@app.route("/api/shares/cached")
def api_shares_cached():
    cache_file = DATA_DIR / "shares_cache.json"
    if cache_file.exists():
        with open(cache_file) as f:
            return jsonify(json.load(f))
    return jsonify({"shares": [], "volumes": {}, "scanned_at": None})


# ---------------------------------------------------------------------------
# Excel parsing / creation
# ---------------------------------------------------------------------------

def _json_safe(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return v


def parse_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw:
        return {"headers": [], "rows": []}

    headers = [str(c).strip() if c is not None else "" for c in rows_raw[0]]

    rows = []
    for raw in rows_raw[1:]:
        if all(c is None for c in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            row[h] = _json_safe(raw[i] if i < len(raw) else None)
        rows.append(row)

    return {"headers": headers, "rows": rows}


def _find_col(headers: list, *keywords) -> Optional[str]:
    """Return first header that contains any of the keywords (case-insensitive)."""
    for h in headers:
        hl = h.lower()
        if any(k in hl for k in keywords):
            return h
    return None


def build_excel(data: dict, cfg: Optional[dict] = None) -> openpyxl.Workbook:
    headers    = data.get("headers", [])
    rows       = data.get("rows", [])
    mailbox_gb = (cfg or load_config()).get("mailbox_gb", 10)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opslag"

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold   = Font(bold=True)
    center = Alignment(horizontal="center")

    # Write headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font  = bold
        c.fill  = yellow
        c.alignment = center

    # Identify special columns for formula reconstruction
    mailbox_h  = _find_col(headers, "mailbox")
    gebruik_h  = _find_col(headers, "gebruik", "used storage")
    factuur_h  = _find_col(headers, "factuur", "factureren", "invoice")

    def hdr_letter(h):
        if h is None:
            return None
        try:
            idx = headers.index(h) + 1
            return get_column_letter(idx)
        except ValueError:
            return None

    m_letter = hdr_letter(mailbox_h)
    g_letter = hdr_letter(gebruik_h)
    f_idx    = (headers.index(factuur_h) + 1) if factuur_h and factuur_h in headers else None

    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            # Re-insert billing formula if all columns are known
            if f_idx == ci and m_letter and g_letter:
                ws.cell(row=ri, column=ci,
                        value=f"={g_letter}{ri}-({mailbox_gb}*{m_letter}{ri})")
            else:
                v = row.get(h)
                ws.cell(row=ri, column=ci, value=v)

    # Auto-width
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    return wb

# ---------------------------------------------------------------------------
# Version / history helpers
# ---------------------------------------------------------------------------

def _ts() -> str:
    return datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")


def _snapshot_current(reason: str = "pre-save"):
    """Save a snapshot of current.json to edits/ before overwriting."""
    if not CURRENT_FILE.exists():
        return
    with open(CURRENT_FILE) as f:
        old = json.load(f)
    old["_snap_reason"] = reason
    old["_snap_at"]     = datetime.datetime.now().isoformat()
    snap = EDITS_DIR / f"{_ts()}.json"
    with open(snap, "w") as f:
        json.dump(old, f, indent=2, default=str)


def _write_current(data: dict):
    _snapshot_current()
    with open(CURRENT_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)


def _apply_retention(directory: Path, pattern: str, keep: int):
    files = sorted(directory.glob(pattern),
                   key=lambda x: x.stat().st_mtime, reverse=True)
    for f in files[keep:]:
        try:
            f.unlink()
        except OSError:
            pass

# ---------------------------------------------------------------------------
# Excel API routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory("static/html", "index.html")


@app.route("/login")
def login_page():
    return send_from_directory("static/html", "login.html")


@app.route("/static/branding/<path:filename>")
def branding_asset(filename: str):
    if BRANDING_DIR and (BRANDING_DIR / filename).is_file():
        return send_from_directory(str(BRANDING_DIR), filename)
    return send_from_directory("static/assets", filename)


@app.route("/favicon.ico")
def favicon():
    if BRANDING_DIR and (BRANDING_DIR / "favicon.ico").is_file():
        return send_from_directory(str(BRANDING_DIR), "favicon.ico")
    return send_from_directory("static/assets", "favicon.ico")


@app.route("/api/excel/current")
def api_excel_current():
    if not CURRENT_FILE.exists():
        return jsonify({"headers": [], "rows": [], "_meta": {}})
    with open(CURRENT_FILE) as f:
        return jsonify(json.load(f))


@app.route("/api/excel/upload", methods=["POST"])
def api_excel_upload():
    if "file" not in request.files:
        return jsonify({"error": "Geen bestand opgegeven"}), 400
    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Alleen .xlsx of .xls bestanden zijn toegestaan"}), 400

    cfg      = load_config()
    orig     = file.filename
    savename = f"{_ts()}_{orig}"
    dest     = UPLOADS_DIR / savename
    file.save(str(dest))

    try:
        data = parse_excel(str(dest))
        # Apply stored mappings and compute diff before saving
        mapping_diff = compute_mapping_diff(data["headers"], data["rows"], load_mappings())
        data["_meta"] = {
            "source":            "upload",
            "original_filename": orig,
            "upload_id":         savename,
            "uploaded_at":       datetime.datetime.now().isoformat(),
            "row_count":         len(data["rows"]),
        }
        _write_current(data)
        _apply_retention(UPLOADS_DIR, "*.xlsx", cfg["upload_retention"])
        _apply_retention(UPLOADS_DIR, "*.xls",  cfg["upload_retention"])
        _apply_retention(EDITS_DIR,   "*.json", cfg["edit_retention"])
        return jsonify({"success": True, "data": data, "mapping_diff": mapping_diff})
    except Exception as e:
        try:
            dest.unlink()
        except OSError:
            pass
        return jsonify({"error": str(e)}), 500


@app.route("/api/excel/save", methods=["POST"])
def api_excel_save():
    body = request.get_json()
    if not body:
        return jsonify({"error": "Geen data ontvangen"}), 400

    cfg = load_config()
    body.setdefault("_meta", {})
    body["_meta"].update({
        "source":    "edit",
        "saved_at":  datetime.datetime.now().isoformat(),
        "row_count": len(body.get("rows", [])),
    })
    _write_current(body)
    _apply_retention(EDITS_DIR, "*.json", cfg["edit_retention"])
    return jsonify({"success": True})


@app.route("/api/excel/export")
def api_excel_export():
    if not CURRENT_FILE.exists():
        return jsonify({"error": "Geen data om te exporteren"}), 404
    with open(CURRENT_FILE) as f:
        data = json.load(f)

    wb  = build_excel(data, load_config())
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"opslag_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )

# ---------------------------------------------------------------------------
# History API routes
# ---------------------------------------------------------------------------

@app.route("/api/history")
def api_history():
    uploads = []
    for f in sorted(UPLOADS_DIR.glob("*"), key=lambda x: x.stat().st_mtime, reverse=True):
        uploads.append({
            "id":       f.stem,
            "filename": f.name,
            "size":     f.stat().st_size,
            "modified": datetime.datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            "type":     "upload",
        })

    edits = []
    for f in sorted(EDITS_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
        meta = {}
        try:
            with open(f) as fp:
                d = fp.read(4096)           # read just enough for metadata
                d = json.loads(d)
                meta = d.get("_meta", {})
                meta["_snap_reason"] = d.get("_snap_reason", "")
                meta["row_count"]    = len(d.get("rows", []))
        except Exception:
            pass
        edits.append({
            "id":       f.stem,
            "filename": f.name,
            "modified": datetime.datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            "meta":     meta,
            "type":     "edit_snapshot",
        })

    return jsonify({"uploads": uploads, "edits": edits})


@app.route("/api/history/restore/upload/<upload_id>", methods=["POST"])
def api_restore_upload(upload_id):
    upload_id = Path(upload_id).name          # prevent path traversal
    matches = list(UPLOADS_DIR.glob(f"{upload_id}*"))
    if not matches:
        return jsonify({"error": "Upload niet gevonden"}), 404

    cfg = load_config()
    try:
        data = parse_excel(str(matches[0]))
        data["_meta"] = {
            "source":       "restore_from_upload",
            "restored_from": upload_id,
            "restored_at":  datetime.datetime.now().isoformat(),
            "row_count":    len(data["rows"]),
        }
        _write_current(data)
        _apply_retention(EDITS_DIR, "*.json", cfg["edit_retention"])
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/history/restore/edit/<edit_id>", methods=["POST"])
def api_restore_edit(edit_id):
    edit_id   = Path(edit_id).name
    edit_path = EDITS_DIR / f"{edit_id}.json"
    if not edit_path.exists():
        return jsonify({"error": "Momentopname niet gevonden"}), 404

    cfg = load_config()
    with open(edit_path) as f:
        data = json.load(f)

    data.setdefault("_meta", {})
    data["_meta"].update({
        "source":       "restore_from_edit",
        "restored_from": edit_id,
        "restored_at":  datetime.datetime.now().isoformat(),
    })
    _write_current(data)
    _apply_retention(EDITS_DIR, "*.json", cfg["edit_retention"])
    return jsonify({"success": True, "data": data})

# ---------------------------------------------------------------------------
# Mappings API routes
# ---------------------------------------------------------------------------

@app.route("/api/mappings", methods=["GET"])
def api_mappings_get():
    return jsonify(load_mappings())


@app.route("/api/mappings/save", methods=["POST"])
def api_mappings_save():
    body = request.get_json()
    if not body:
        return jsonify({"error": "Geen data"}), 400

    mappings = load_mappings()
    m = dict(mappings.get("map", {}))

    # Apply updates (new or changed mappings) — keys are always lowercase
    for entry in body.get("updates", []):
        key = str(entry.get("key", "")).strip().lower()
        if not key:
            continue
        if entry.get("share"):
            m[key] = {"share": entry["share"], "name": entry.get("name", "")}
        elif key in m:
            # Share explicitly cleared — remove the share but keep the entry as unmapped
            m[key] = {"share": None, "name": entry.get("name", m[key].get("name", ""))}

    # Remove entries user explicitly discarded
    for key in body.get("remove", []):
        m.pop(str(key), None)

    mappings["map"]        = m
    mappings["key_col"]    = body.get("key_col", mappings.get("key_col"))
    mappings["updated_at"] = datetime.datetime.now().isoformat()
    save_mappings(mappings)
    return jsonify({"success": True, "mappings": mappings})


@app.route("/api/mappings/history")
def api_mappings_history():
    snaps = []
    for f in sorted(MAPPINGS_DIR.glob("*.json"),
                    key=lambda x: x.stat().st_mtime, reverse=True):
        try:
            with open(f) as fp:
                d = json.load(fp)
            snaps.append({
                "id":       f.stem,
                "filename": f.name,
                "modified": datetime.datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                "count":    len(d.get("map", {})),
                "snap_at":  d.get("_snap_at", ""),
            })
        except Exception:
            pass
    return jsonify({"snapshots": snaps})


@app.route("/api/mappings/restore/<snap_id>", methods=["POST"])
def api_mappings_restore(snap_id):
    snap_id   = Path(snap_id).name
    snap_path = MAPPINGS_DIR / f"{snap_id}.json"
    if not snap_path.exists():
        return jsonify({"error": "Momentopname niet gevonden"}), 404
    with open(snap_path) as f:
        data = json.load(f)
    save_mappings(data)
    return jsonify({"success": True, "mappings": data})


# ---------------------------------------------------------------------------
# Settings API routes
# ---------------------------------------------------------------------------

@app.route("/api/settings", methods=["GET"])
def api_settings_get():
    cfg = load_config()
    cfg.pop('auth_password', None)
    cfg.pop('dsm_password', None)
    cfg['dsm_password_set'] = bool(load_config().get('dsm_password'))
    return jsonify(cfg)


@app.route("/api/settings", methods=["POST"])
def api_settings_post():
    body = request.get_json()
    if not body:
        return jsonify({"error": "Geen data ontvangen"}), 400

    cfg = load_config()

    for key in ("share_paths", "exclude_shares"):
        if key in body and isinstance(body[key], list):
            cfg[key] = [str(v).strip() for v in body[key] if str(v).strip()]

    for key in ("upload_retention", "edit_retention"):
        if key in body:
            val = int(body[key])
            if 1 <= val <= 100:
                cfg[key] = val

    if "mailbox_gb" in body:
        val = float(body["mailbox_gb"])
        if val >= 0:
            cfg["mailbox_gb"] = val

    if "auth_enabled" in body:
        cfg["auth_enabled"] = bool(body["auth_enabled"])
    if "auth_username" in body and str(body["auth_username"]).strip():
        cfg["auth_username"] = str(body["auth_username"]).strip()
    if "auth_password" in body and str(body["auth_password"]).strip():
        cfg["auth_password"] = hash_password(str(body["auth_password"]).strip())

    for key in ("dsm_host", "dsm_user"):
        if key in body:
            cfg[key] = str(body[key]).strip()
    if "dsm_port" in body:
        try:
            val = int(body["dsm_port"])
            if 1 <= val <= 65535:
                cfg["dsm_port"] = val
        except (ValueError, TypeError):
            pass
    if "dsm_password" in body and str(body["dsm_password"]).strip():
        cfg["dsm_password"] = _encrypt_credential(str(body["dsm_password"]).strip())

    save_config(cfg)
    _apply_retention(UPLOADS_DIR, "*.xlsx", cfg["upload_retention"])
    _apply_retention(UPLOADS_DIR, "*.xls",  cfg["upload_retention"])
    _apply_retention(EDITS_DIR,   "*.json", cfg["edit_retention"])
    cfg.pop('auth_password', None)
    return jsonify({"success": True, "config": cfg})

@app.route("/api/settings/test_dsm", methods=["POST"])
def api_test_dsm():
    body     = request.get_json() or {}
    host     = str(body.get('dsm_host', '')).strip()
    user     = str(body.get('dsm_user', '')).strip()
    password = str(body.get('dsm_password', '')).strip()
    try:
        port = int(body.get('dsm_port', 3333))
    except (ValueError, TypeError):
        port = 3333

    # Frontend may omit password when using the already-stored one
    if not password and body.get('use_stored_password'):
        password = _decrypt_credential(str(load_config().get('dsm_password', '')).strip())

    if not host or not user or not password:
        return jsonify({'error': 'Vul host, gebruiker en wachtwoord in'}), 400

    base_url = f"http://{host}:{port}"
    cj       = http.cookiejar.CookieJar()
    opener   = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))

    try:
        login_data = urllib.parse.urlencode({
            'api': 'SYNO.API.Auth', 'version': '3', 'method': 'login',
            'account': user, 'passwd': password,
            'session': 'SynologyToolTest', 'format': 'cookie',
        })
        resp   = opener.open(f"{base_url}/webapi/entry.cgi", login_data.encode(), timeout=10)
        result = json.loads(resp.read())
        if not result.get('success'):
            code = result.get('error', {}).get('code', '?')
            return jsonify({'error': f'DSM login mislukt (code {code})'}), 400

        list_data = urllib.parse.urlencode({
            'api': 'SYNO.Core.Report', 'version': '1', 'method': 'list',
        })
        resp    = opener.open(f"{base_url}/webapi/entry.cgi", list_data.encode(), timeout=15)
        result  = json.loads(resp.read())
        if not result.get('success'):
            return jsonify({'error': 'Login gelukt maar Storage Analyzer API niet beschikbaar'}), 400

        reports = result['data'].get('reports', [])
        return jsonify({'success': True, 'report_count': len(reports)})
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500


# ---------------------------------------------------------------------------
# DSM monthly reports setup
# ---------------------------------------------------------------------------

@app.route("/api/dsm/setup_monthly_reports", methods=["POST"])
def api_dsm_setup_monthly_reports():
    cfg      = load_config()
    dsm_user = cfg.get('dsm_user', '').strip()
    dsm_pw_enc = cfg.get('dsm_password', '').strip()
    dsm_host = cfg.get('dsm_host', 'localhost')
    dsm_port = int(cfg.get('dsm_port', 3333))

    if not dsm_user or not dsm_pw_enc:
        return jsonify({'error': 'DSM-inloggegevens niet ingesteld in Instellingen'}), 400

    body   = request.get_json() or {}
    sched_day    = max(1, min(28, int(body.get('day',    1))))
    sched_hour   = max(0, min(23, int(body.get('hour',   3))))
    sched_minute = max(0, min(59, int(body.get('minute', 0))))
    requested_shares = [str(s).strip() for s in body.get('shares', []) if str(s).strip()]

    dsm_pw   = _decrypt_credential(dsm_pw_enc)
    base_url = f"http://{dsm_host}:{dsm_port}"
    exclude  = set(cfg.get('exclude_shares', []))

    cj     = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))

    def dsm_post(params):
        data = urllib.parse.urlencode(params)
        resp = opener.open(f"{base_url}/webapi/entry.cgi", data.encode(), timeout=15)
        return json.loads(resp.read())

    # Authenticate
    try:
        r = dsm_post({
            'api': 'SYNO.API.Auth', 'version': '3', 'method': 'login',
            'account': dsm_user, 'passwd': dsm_pw,
            'session': 'SynologyToolSetup', 'format': 'cookie',
        })
        if not r.get('success'):
            code = r.get('error', {}).get('code', '?')
            return jsonify({'error': f'DSM login mislukt (code {code})'}), 400
    except Exception as ex:
        return jsonify({'error': f'DSM verbindingsfout: {ex}'}), 500

    out = {
        'existing_reports': [],
        'covered_shares':   [],
        'created':          [],
        'failed':           [],
        'schedule':         {},
        'schedule_type':    None,
        'schedule_set':     False,
        'task_created':     False,
        'errors':           [],
    }

    # List existing reports
    try:
        r       = dsm_post({'api': 'SYNO.Core.Report', 'version': '1', 'method': 'list'})
        reports = r.get('data', {}).get('reports', []) if r.get('success') else []
    except Exception as ex:
        reports = []
        out['errors'].append(f'Rapportlijst ophalen mislukt: {ex}')

    covered = set()
    for rpt in reports:
        rname = rpt.get('name', '')
        if rname:
            out['existing_reports'].append(rname)
        for sh in rpt.get('shares', []):
            if isinstance(sh, str):
                covered.add(sh)
            elif isinstance(sh, dict):
                covered.add(sh.get('name', '') or sh.get('share_name', ''))

    # Get current schedule config
    try:
        r = dsm_post({'api': 'SYNO.Core.Report.Config', 'version': '1', 'method': 'get'})
        if r.get('success'):
            out['schedule'] = r.get('data', {})
    except Exception:
        pass

    report_location = out['schedule'].get('report_location', '')

    # Discover non-filtered shares from configured share paths
    non_filtered = set()
    for sp in cfg.get('share_paths', ['/volume1']):
        try:
            for entry in os.listdir(sp):
                if entry not in exclude and not entry.startswith('@') and not entry.startswith('#'):
                    non_filtered.add(entry)
        except Exception:
            pass

    out['covered_shares'] = sorted(covered)

    # Use explicitly requested shares if provided; otherwise auto-discover all non-filtered
    if requested_shares:
        uncovered = [s for s in requested_shares if s not in covered]
    else:
        uncovered = sorted(non_filtered - covered)

    # Create Storage Analyzer report profiles for uncovered shares
    for share in uncovered:
        # Use a safe, unique id derived from share name (lowercase, underscores)
        report_id = 'syntool_' + re.sub(r'[^a-z0-9_]', '_', share.lower())
        try:
            r = dsm_post({
                'api': 'SYNO.Core.Report', 'version': '1', 'method': 'create',
                'id': report_id,
                'shares[]': share,
            })
            if r.get('success'):
                out['created'].append(share)
                covered.add(share)
            else:
                code = r.get('error', {}).get('code', '?')
                msg  = r.get('error', {}).get('errors', {}).get('msg', '')
                if code == 4907:
                    # Folder already exists — report already set up under a different name
                    out['covered_shares'].append(share)
                    covered.add(share)
                else:
                    out['failed'].append({'share': share, 'code': code, 'msg': msg})
        except Exception as ex:
            out['failed'].append({'share': share, 'error': str(ex)})

    # Set schedule via Storage Analyzer config — only include report_location when set.
    # Monthly is not natively supported; weekly is the best this API offers.
    schedule_set = False
    h, m, d = str(sched_hour), str(sched_minute), str(sched_day)
    base_params = {'hour': h, 'minute': m}
    if report_location:
        base_params['report_location'] = report_location
    for extra, label in [
        ({'schedule_type': 'monthly', 'month_day': d}, 'monthly'),
        ({'month_day': d}, 'monthly'),
        ({'week_day': '1'}, 'weekly_monday'),
    ]:
        if schedule_set:
            break
        try:
            r = dsm_post(dict(api='SYNO.Core.Report.Config', version='1', method='set',
                              **base_params, **extra))
            if r.get('success'):
                schedule_set = True
                out['schedule_type'] = label
        except Exception:
            pass

    # Fallback: DSM Task Scheduler script task.
    # Delete any pre-existing task with the same name first (avoids 4800 name-conflict).
    # Try API versions 1 then 4 — version requirement varies by DSM release.
    if not schedule_set:
        cmd = '/usr/syno/bin/syno_volume_analyze -w eval-timetable'
        task_name = 'Storage Analyzer Maandelijks'

        # Delete existing task with this name if present
        try:
            r = dsm_post({'api': 'SYNO.Core.TaskScheduler', 'version': '1', 'method': 'list'})
            tasks = r.get('data', {}).get('tasks', []) if r.get('success') else []
            for t in tasks:
                if t.get('name') == task_name and t.get('can_delete'):
                    dsm_post({'api': 'SYNO.Core.TaskScheduler', 'version': '1',
                              'method': 'delete', 'id': str(t['id'])})
        except Exception:
            pass

        schedule_json = json.dumps({
            'date': str(sched_day), 'date_type': 0,
            'hour': sched_hour, 'minute': sched_minute,
            'repeat_min': 0, 'repeat_min_store_config': 30,
            'week_day': '0,1,2,3,4,5,6',
        })
        last_error = ''
        for ts_ver in ['1', '4']:
            if schedule_set:
                break
            task_json = json.dumps({'owner': 'root', 'script': cmd})
            try:
                r = dsm_post({
                    'api': 'SYNO.Core.TaskScheduler', 'version': ts_ver, 'method': 'create',
                    'name': task_name, 'owner': 'root', 'enable': 'true',
                    'schedule': schedule_json, 'task': task_json,
                })
                if r.get('success'):
                    schedule_set = True
                    out['schedule_type'] = 'task_scheduler_monthly'
                    out['task_created']  = True
                else:
                    code = r.get('error', {}).get('code', '?')
                    last_error = f'code {code} (v{ts_ver})'
            except Exception as ex:
                last_error = str(ex)

        if not schedule_set:
            out['errors'].append(f'Task Scheduler aanmaken mislukt ({last_error}) — stel handmatig in via DSM')

    out['schedule_set']     = schedule_set
    out['covered_shares']   = sorted(covered)
    return jsonify({'success': True, **out})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    ensure_dirs()
    port  = int(os.environ.get("PORT", 9000))
    host  = os.environ.get("HOST", "0.0.0.0")
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    print(f"Synology Storage Tool  →  http://{host}:{port}")
    print(f"Data directory         →  {DATA_DIR.resolve()}")
    app.run(host=host, port=port, debug=debug, threaded=True)
